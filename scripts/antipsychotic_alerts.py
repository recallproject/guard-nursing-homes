#!/usr/bin/env python3
"""
Antipsychotic Prescribing Pattern Analysis
===========================================
Combines MDS Measure 481 (antipsychotic use rate), LTCFocus schizophrenia/bipolar
diagnosis rates, chemical restraint flags, prescriber concentration, and
YOY trends to produce a review priority score for each nursing facility.

Output: antipsychotic_alerts.json with facilities scoring >= 3.
"""

import csv
import json
import os
import sys
from collections import defaultdict

import openpyxl

# ── File paths ──────────────────────────────────────────────────────────
MDS_QM = "/Volumes/LaCie/RB7-Project_RECALL-Missing-Persons/healthcare_fraud/nursing_home/data/mds_quality_measures.csv"
CHEM_FLAGS = "/Volumes/LaCie/RB7-Project_RECALL-Missing-Persons/healthcare_fraud/dmepos/antipsychotic_chemical_restraint_flags.csv"
SNF_MATCHES = "/Volumes/LaCie/RB7-Project_RECALL-Missing-Persons/healthcare_fraud/dmepos/antipsychotic_snf_matches.csv"
YOY_FAC = "/Volumes/LaCie/RB7-Project_RECALL-Missing-Persons/healthcare_fraud/dmepos/antipsychotic_yoy_facilities.csv"
LTCFOCUS_2023 = os.path.expanduser("~/Desktop/ltcfocus_brown/facility_2023_A.xlsx")
LTCFOCUS_2022 = os.path.expanduser("~/Desktop/ltcfocus_brown/facility_2022.xlsx")
OUTPUT = "/Users/rob/Projects/oversight-reports/public/data/antipsychotic_alerts.json"

NATIONAL_AVG_AP = 9.3  # CMS national average antipsychotic rate (%)


def safe_float(val, default=None):
    """Convert to float, returning default on failure."""
    if val is None:
        return default
    s = str(val).strip()
    if s in ("", "LNE", "None", "nan"):
        return default
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


# ═══════════════════════════════════════════════════════════════════════
# STEP 1: Load MDS Measure 481 (antipsychotic use rate per facility)
# ═══════════════════════════════════════════════════════════════════════
print("STEP 1: Loading MDS Measure 481 data...")

measure481 = {}  # CCN -> antipsychotic rate (Four Quarter Average)
facility_info = {}  # CCN -> {name, state, ...}

with open(MDS_QM, "r", encoding="latin-1") as f:
    reader = csv.DictReader(f)
    for row in reader:
        if row["Measure Code"] == "481":
            ccn = row["CMS Certification Number (CCN)"].strip()
            rate = safe_float(row["Four Quarter Average Score"])
            if rate is not None:
                measure481[ccn] = rate
                facility_info[ccn] = {
                    "name": row["Provider Name"].strip(),
                    "state": row["State"].strip(),
                    "city": row["City/Town"].strip(),
                    "zip": row["ZIP Code"].strip(),
                }

print(f"  Measure 481 loaded: {len(measure481):,} facilities with valid scores")
above_nat = sum(1 for v in measure481.values() if v > NATIONAL_AVG_AP)
print(f"  Above national avg ({NATIONAL_AVG_AP}%): {above_nat:,}")


# ═══════════════════════════════════════════════════════════════════════
# STEP 2: Load LTCFocus schizophrenia/bipolar diagnosis rates
# ═══════════════════════════════════════════════════════════════════════
print("\nSTEP 2: Loading LTCFocus schizophrenia/bipolar diagnosis data...")


def load_ltcfocus_schiz(filepath, year_label):
    """Load pctschiz_bipol from an LTCFocus facility Excel file."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    ccn_idx = headers.index("PROV1680")
    state_idx = headers.index("state")
    schiz_idx = headers.index("pctschiz_bipol")

    data = {}  # CCN -> rate
    for row in ws.iter_rows(min_row=2, values_only=True):
        ccn = row[ccn_idx]
        if not ccn:
            continue
        ccn = str(ccn).strip()
        rate = safe_float(row[schiz_idx])
        if rate is not None:
            data[ccn] = {"rate": rate, "state": str(row[state_idx]).strip() if row[state_idx] else ""}
    wb.close()
    print(f"  {year_label}: {len(data):,} facilities with valid schiz/bipolar data")
    return data


schiz_2023 = load_ltcfocus_schiz(LTCFOCUS_2023, "2023")
schiz_2022 = load_ltcfocus_schiz(LTCFOCUS_2022, "2022")

# Compute state averages for 2023
state_schiz_vals = defaultdict(list)
for ccn, info in schiz_2023.items():
    state_schiz_vals[info["state"]].append(info["rate"])

state_schiz_avg = {st: sum(vals) / len(vals) for st, vals in state_schiz_vals.items() if vals}
print(f"  State averages computed for {len(state_schiz_avg)} states")


# ═══════════════════════════════════════════════════════════════════════
# STEP 3: Load chemical restraint flags
# ═══════════════════════════════════════════════════════════════════════
print("\nSTEP 3: Loading chemical restraint flags...")

chem_flags = {}  # CCN -> True/False
with open(CHEM_FLAGS, "r") as f:
    reader = csv.DictReader(f)
    for row in reader:
        ccn = row["CCN"].strip()
        chem_flags[ccn] = row["chemical_restraint_flag"].strip() == "True"

flagged_count = sum(1 for v in chem_flags.values() if v)
print(f"  Chemical restraint data: {len(chem_flags):,} facilities, {flagged_count:,} flagged")


# ═══════════════════════════════════════════════════════════════════════
# STEP 4: Load prescriber concentration
# ═══════════════════════════════════════════════════════════════════════
print("\nSTEP 4: Computing prescriber concentration...")

fac_prescriber_claims = defaultdict(lambda: defaultdict(float))
with open(SNF_MATCHES, "r") as f:
    reader = csv.DictReader(f)
    for row in reader:
        ccn = row["CCN"].strip()
        npi = row["NPI"].strip()
        claims = safe_float(row["total_ap_claims"], 0)
        fac_prescriber_claims[ccn][npi] += claims

prescriber_concentration = {}  # CCN -> max prescriber share (0-1)
for ccn, prescribers in fac_prescriber_claims.items():
    total = sum(prescribers.values())
    if total > 0:
        max_share = max(prescribers.values()) / total
        prescriber_concentration[ccn] = max_share

high_conc = sum(1 for v in prescriber_concentration.values() if v > 0.5)
print(f"  Prescriber concentration: {len(prescriber_concentration):,} facilities")
print(f"  >50% from one prescriber: {high_conc:,}")


# ═══════════════════════════════════════════════════════════════════════
# STEP 5: Load YOY antipsychotic trends
# ═══════════════════════════════════════════════════════════════════════
print("\nSTEP 5: Loading YOY antipsychotic trends...")

yoy_data = {}  # CCN -> {claims_change, claims_pct_change, ...}
with open(YOY_FAC, "r") as f:
    reader = csv.DictReader(f)
    for row in reader:
        ccn = row["CCN"].strip()
        claims_2022 = safe_float(row.get("total_claims_2022"), 0)
        claims_2023 = safe_float(row.get("total_claims_2023"), 0)
        pct_change = safe_float(row.get("claims_pct_change"))
        yoy_data[ccn] = {
            "claims_2022": claims_2022,
            "claims_2023": claims_2023,
            "pct_change": pct_change,
        }

print(f"  YOY data: {len(yoy_data):,} facilities")


# ═══════════════════════════════════════════════════════════════════════
# STEP 6: Compute review priority scores
# ═══════════════════════════════════════════════════════════════════════
print("\nSTEP 6: Computing review priority scores...")

# Build universe of all CCNs across data sources
all_ccns = set()
all_ccns.update(measure481.keys())
all_ccns.update(chem_flags.keys())
all_ccns.update(yoy_data.keys())
all_ccns.update(schiz_2023.keys())

print(f"  Total unique CCNs across all sources: {len(all_ccns):,}")

results = {}
score_distribution = defaultdict(int)
data_issues = []

for ccn in all_ccns:
    score = 0
    factors = []

    # --- Antipsychotic rate (Measure 481) ---
    ap_rate = measure481.get(ccn)
    if ap_rate is None:
        continue  # No Measure 481 = can't score meaningfully

    if ap_rate > 30:
        score += 5
        factors.append(f"Antipsychotic rate {ap_rate:.1f}% (>30%, extremely high)")
    elif ap_rate > 20:
        score += 3
        factors.append(f"Antipsychotic rate {ap_rate:.1f}% (>20%, very high)")
    elif ap_rate > 15:
        score += 2
        factors.append(f"Antipsychotic rate {ap_rate:.1f}% (>15%, high)")

    # --- Chemical restraint flag ---
    is_chem_flagged = chem_flags.get(ccn, False)
    if is_chem_flagged:
        score += 3
        factors.append("Chemical restraint flag (low RN staffing + high AP prescribing)")

    # --- YOY antipsychotic trend ---
    yoy = yoy_data.get(ccn)
    yoy_trend = "unknown"
    if yoy and yoy["pct_change"] is not None:
        if yoy["pct_change"] > 10:
            yoy_trend = "increasing"
            score += 1
            factors.append(f"YOY antipsychotic claims increased {yoy['pct_change']:.0f}%")
        elif yoy["pct_change"] < -10:
            yoy_trend = "decreasing"
        else:
            yoy_trend = "stable"

    # --- Schizophrenia/bipolar diagnosis pattern signal ---
    schiz_info = schiz_2023.get(ccn)
    schiz_dx_rate = None
    schiz_state_avg = None
    schiz_yoy_change = None

    if schiz_info:
        schiz_dx_rate = schiz_info["rate"]
        state = schiz_info["state"]
        schiz_state_avg = state_schiz_avg.get(state)

        # YOY change in schizophrenia dx rate
        schiz_2022_info = schiz_2022.get(ccn)
        if schiz_2022_info:
            schiz_yoy_change = schiz_dx_rate - schiz_2022_info["rate"]

        # Diagnostic pattern signal: facility schiz rate significantly above state avg AND high AP use
        if schiz_state_avg and schiz_dx_rate > (schiz_state_avg * 1.5) and ap_rate > NATIONAL_AVG_AP:
            score += 4
            factors.append(
                f"Schizophrenia/bipolar dx rate {schiz_dx_rate:.1f}% vs state avg {schiz_state_avg:.1f}% "
                f"(>{schiz_state_avg * 1.5:.1f}% threshold) with above-avg AP use"
            )

        # Extra signal: schiz rate increasing YOY while already above state avg
        if schiz_yoy_change is not None and schiz_yoy_change > 3 and schiz_state_avg and schiz_dx_rate > schiz_state_avg:
            score += 1
            factors.append(
                f"Schizophrenia/bipolar dx rate increased {schiz_yoy_change:+.1f}pp YOY"
            )

    # --- Prescriber concentration ---
    conc = prescriber_concentration.get(ccn)
    if conc is not None and conc > 0.5:
        score += 2
        factors.append(f"Prescriber concentration: {conc * 100:.0f}% of AP claims from one prescriber")

    # --- Classify risk level ---
    if score >= 8:
        risk_level = "critical"
    elif score >= 5:
        risk_level = "high"
    elif score >= 3:
        risk_level = "elevated"
    else:
        risk_level = "normal"

    score_distribution[risk_level] += 1

    if score >= 3:
        entry = {
            "name": facility_info.get(ccn, {}).get("name", ""),
            "state": facility_info.get(ccn, {}).get("state", ""),
            "antipsychotic_rate": round(ap_rate, 1),
            "national_avg": NATIONAL_AVG_AP,
            "risk_score": score,
            "risk_level": risk_level,
            "chemical_restraint_flag": is_chem_flagged,
            "yoy_trend": yoy_trend,
            "factors": factors,
        }

        if schiz_dx_rate is not None:
            entry["schizophrenia_dx_rate"] = round(schiz_dx_rate, 1)
        if schiz_state_avg is not None:
            entry["schizophrenia_state_avg"] = round(schiz_state_avg, 1)
        if schiz_yoy_change is not None:
            entry["schizophrenia_yoy_change"] = round(schiz_yoy_change, 1)

        if conc is not None:
            entry["prescriber_concentration"] = round(conc, 2)

        results[ccn] = entry


# ═══════════════════════════════════════════════════════════════════════
# STEP 7: Write output
# ═══════════════════════════════════════════════════════════════════════
print("\nSTEP 7: Writing output...")

# Sort by risk score descending
sorted_results = dict(
    sorted(results.items(), key=lambda x: x[1]["risk_score"], reverse=True)
)

os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
with open(OUTPUT, "w") as f:
    json.dump(sorted_results, f, indent=2)

file_size = os.path.getsize(OUTPUT)
print(f"  Output: {OUTPUT}")
print(f"  File size: {file_size:,} bytes ({file_size / 1024 / 1024:.1f} MB)")

# ═══════════════════════════════════════════════════════════════════════
# REPORT
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("ANTIPSYCHOTIC PRESCRIBING PATTERN ANALYSIS — REPORT")
print("=" * 60)

print(f"\nData sources loaded:")
print(f"  MDS Measure 481:       {len(measure481):>6,} facilities")
print(f"  LTCFocus 2023 schiz:   {len(schiz_2023):>6,} facilities")
print(f"  LTCFocus 2022 schiz:   {len(schiz_2022):>6,} facilities")
print(f"  Chemical restraint:    {len(chem_flags):>6,} facilities")
print(f"  YOY data:              {len(yoy_data):>6,} facilities")
print(f"  Prescriber conc:       {len(prescriber_concentration):>6,} facilities")

print(f"\nRisk level distribution (all {len(all_ccns):,} scored facilities):")
for level in ["critical", "high", "elevated", "normal"]:
    count = score_distribution.get(level, 0)
    print(f"  {level.upper():<12s}: {count:>6,}")

print(f"\nOutput: {len(sorted_results):,} facilities with score >= 3")
print(f"File: {OUTPUT}")
print(f"Size: {file_size:,} bytes ({file_size / 1024 / 1024:.1f} MB)")

# Top 20 highest risk
print(f"\nTop 20 highest risk facilities:")
for i, (ccn, entry) in enumerate(sorted_results.items()):
    if i >= 20:
        break
    print(
        f"  {ccn}  {entry['name'][:40]:<40s}  {entry['state']:<3s}  "
        f"Score={entry['risk_score']:>2}  AP={entry['antipsychotic_rate']:>5.1f}%  "
        f"Level={entry['risk_level']}"
    )
    for factor in entry["factors"]:
        print(f"         - {factor}")

# Data issues
if len(measure481) < 10000:
    print(f"\n  WARNING: Only {len(measure481):,} facilities had Measure 481 data (expected ~14,700)")
schiz_overlap = len(set(measure481.keys()) & set(schiz_2023.keys()))
print(f"\n  Measure 481 + schiz data overlap: {schiz_overlap:,} facilities")
chem_overlap = len(set(measure481.keys()) & set(chem_flags.keys()))
print(f"  Measure 481 + chemical restraint overlap: {chem_overlap:,} facilities")

print("\nDone.")
